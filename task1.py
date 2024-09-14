import pybedtools
import requests
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re

class TargetRegions:
    def __init__(self, bed_file_path, gtf_file_path, intersected_file_path='intersected.bed', uniprot_excel_path='uniprot_protein_data.xlsx', debug = False):
        self.bed_file_path = bed_file_path
        self.gtf_file_path = gtf_file_path
        self.intersected_file_path = 'intersected_results/' + intersected_file_path
        self.uniprot_excel_path = 'gene_disease_info/' + uniprot_excel_path
        self.debug = debug
        self.df_bed = None
        self.metadata_line = ''
        self.df_exons = None
        self.df_bed_final = None

    def debug_log(self, data, is_important=False, is_error=False):
        if self.debug:
            if is_error:
                print(f"Произошла ошибка: {data}")
            else:
                print(f"Данные логирования: {data}")
        elif is_important:
            print(f"Произошла критическая ошибка: {data}")
            raise

    def add_gene_id_and_exon(self):
        try:
            self.ensure_directories_exist()
            self.read_metadata()
            self.intersect_files()
            self.load_data()
            self.process_exons()
            self.merge_and_save()

            print(f'В файл {self.bed_file_path} были добавлены столбцы с id гена и номером экзона.')

        except Exception as e:
            self.debug_log(e, is_important=True, is_error=True)

    def ensure_directories_exist(self):
        os.makedirs('intersected_results', exist_ok=True)
        os.makedirs('gene_disease_info', exist_ok=True)

    def get_prot_info(self):
        try:
            self.create_prot_excel()
        except Exception as e:
            self.debug_log(e, is_important=True, is_error=True)

    def read_metadata(self):
        with open(self.bed_file_path, 'r') as bed_file:
            first_line = bed_file.readline().strip()
            if first_line.startswith('track'):
                self.metadata_line = first_line
            else:
                self.metadata_line = None

    def intersect_files(self):
        bed_file = pybedtools.BedTool(self.bed_file_path)
        gtf_file = pybedtools.BedTool(self.gtf_file_path)
        intersected = bed_file.intersect(gtf_file, wa=True, wb=True)
        intersected.saveas(self.intersected_file_path)

    def load_data(self):
        intersected = pybedtools.BedTool(self.intersected_file_path)
        df_intesected = intersected.to_dataframe(header=None, names=[
            'chr', 'start_ampl', 'end_ampl', 'name_ampl', 'score', 'pool',
            'chr_2', 'source', 'location', 'start2', 'end2', 'score2', 'strand', '.', 'attributes'
        ])
        self.df_exons = df_intesected[df_intesected['location'] == 'exon'].copy()
        self.df_exons = self.df_exons.drop_duplicates(subset=['name_ampl'])

        self.df_bed = pd.read_csv(self.bed_file_path, sep='\t', header=0, names=[
            'chr', 'start', 'end', 'name_ampl', 'score', 'pool'
        ])

    @staticmethod
    def extract_attribute(attribute_str, key):
        pattern = rf'{key} "([^"]+)"'
        match = re.search(pattern, attribute_str)
        return match.group(1) if match else None

    def process_exons(self):
        self.df_exons['gene_id'] = self.df_exons['attributes'].apply(lambda x: self.extract_attribute(x, 'gene_id'))
        self.df_exons['exon_number'] = self.df_exons['attributes'].apply(
            lambda x: self.extract_attribute(x, 'exon_number'))

    def merge_and_save(self):
        df_merged = pd.merge(self.df_bed, self.df_exons[['name_ampl', 'gene_id', 'exon_number']], on='name_ampl',
                             how='left')
        df_merged['gene_id'] = df_merged['gene_id'].fillna("information wasn't found")
        df_merged['exon_number'] = df_merged['exon_number'].fillna("information wasn't found")

        bed_columns = ['chr', 'start', 'end', 'name_ampl', 'score', 'pool', 'gene_id', 'exon_number']
        self.df_bed_final = df_merged[bed_columns]

        with open(self.bed_file_path, 'w') as bed_file:
            bed_file.write(self.metadata_line + '\n')

            self.df_bed_final.to_csv(bed_file, sep='\t', header=False, index=False)

    def extract_gene_ids(self):
        gene_ids = list(set(self.df_bed_final[self.df_bed_final['gene_id'] != "information wasn't found"]['gene_id'].tolist()))
        return gene_ids

    def fetch_uniprot_data(self, uniprot_ids):
        base_url = "https://rest.uniprot.org/uniprotkb/"
        protein_data = []

        for uniprot_id in uniprot_ids:
            query_url = f"{base_url}{uniprot_id}?format=json"
            try:
                response = requests.get(query_url)
                response.raise_for_status()
                data = response.json()
                protein_name = data.get('proteinDescription', {}).get('recommendedName', {}).get('fullName',{}).get('value','N/A')
                comments = data.get('comments', [])
                diseases = [comment.get('disease', {}).get('diseaseId', '') for comment in comments if
                            comment.get('commentType') == 'DISEASE']
                diseases_str = ', '.join(diseases) if diseases else '-'

                protein_data.append({
                    'UniProt ID': uniprot_id,
                    'Protein Name': protein_name,
                    'Disease Description': diseases_str
                })
                self.debug_log(query_url)

            except requests.exceptions.RequestException as e:
                self.debug_log(e, is_error=True)

        return protein_data

    def create_prot_excel(self):
        if self.df_bed_final is not None:
            gene_ids = self.extract_gene_ids()
            protein_data = self.fetch_uniprot_data(gene_ids)

            df_proteins = pd.DataFrame(protein_data)
            df_proteins.to_excel(self.uniprot_excel_path, index=False, engine='openpyxl')

            output_dir = 'gene_disease_info'
            os.makedirs(output_dir, exist_ok=True)
            excel_path = os.path.join(output_dir, 'uniprot_protein_data.xlsx')

            df_proteins.to_excel(excel_path, index=False, engine='openpyxl')

            # Форматирование excel файла
            wb = load_workbook(excel_path)
            ws = wb.active

            column_widths = {
                'A': 20,  # UniProt ID
                'B': 30,  # Название белка
                'C': 50,  # Связанные заболевания
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            for row in ws.iter_rows(min_row=2, max_col=4):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row[0].row].height = 60

            wb.save(self.uniprot_excel_path)

            print(f"Файл с с информацией об экспрессируемых с генов белках успешно создан: {self.uniprot_excel_path}")

        else:
            print('Последовательности не были определены')


if __name__ == "__main__":
    processor = TargetRegions(
        bed_file_path=os.getenv('BED_FILE_PATH'),
        gtf_file_path=os.getenv('GTF_FILE_PATH'),
        intersected_file_path=os.getenv('INTERSECTED_PATH'),
        uniprot_excel_path=os.getenv('UNIPROT_INFO'),
        debug=os.getenv('DEBUG')
    )
    processor.add_gene_id_and_exon()
    processor.create_prot_excel()
